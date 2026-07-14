"""从 MeterSphere 拉取模块子计划，并生成与日报模板一致的 Excel 表。"""
from __future__ import annotations

import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]


def run_fetch() -> Path:
    env = dict(**{k: v for k, v in __import__("os").environ.items()})
    env.setdefault("METERSPHERE_MODULE_NAME", "OBIS-20260622-20260703")
    env["PYTHONIOENCODING"] = "utf-8"
    subprocess.check_call(
        [sys.executable, str(ROOT / "scripts" / "ms_fetch_module_plans.py")],
        cwd=str(ROOT),
        env=env,
    )
    files = sorted((ROOT / "exports" / "metersphere").glob("module_*_execution_*.json"))
    if not files:
        raise SystemExit("未找到 execution json，请先确认 ms_fetch_module_plans.py 成功")
    return files[-1]


def load_plans(json_path: Path) -> tuple[dict, list[dict]]:
    import json

    data = json.loads(json_path.read_text(encoding="utf-8"))
    plans = data.get("plans") or []
    # 稳定排序：计划组名 + 计划名
    plans = sorted(plans, key=lambda p: (p.get("parentGroupName") or "", p.get("name") or ""))
    return data, plans


def style_header_row(ws, row: int, fills: dict[str, PatternFill], font: Font, border: Border):
    for col, fill_key in fills.items():
        cell = ws.cell(row, col)
        cell.fill = fill_key
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def build_workbook(module_name: str, summary: dict, plans: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试执行"

    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    yellow = PatternFill("solid", fgColor="FFFF00")
    grey = PatternFill("solid", fgColor="D9D9D9")
    blue = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(bold=True)
    green_font = Font(color="00B050", bold=True)

    # 标题区
    ws["B2"] = f"Sprint_Child_Plans_{datetime.now().strftime('%m/%d')}"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B3"] = "Sprint"
    ws["C3"] = f"【{module_name}】"
    ws["B4"] = "Summary"
    ws["C4"] = (
        f"子计划数：{summary.get('planCount', len(plans))} | "
        f"用例总数：{summary.get('caseTotal', 0)} | "
        f"整体通过率：{summary.get('passRate', '-')}%"
    )

    # 表头两行（对齐日报模板）
    # row 6 category, row 7 sub headers
    ws.merge_cells("B6:B7")
    ws["B6"] = "Story"
    ws.merge_cells("C6:E6")
    ws["C6"] = "Test Design"
    ws.merge_cells("F6:J6")
    ws["F6"] = "Test Execution"

    ws["C7"] = "Design"
    ws["D7"] = "Review"
    ws["E7"] = "Review Rate"
    ws["F7"] = "PASS"
    ws["G7"] = "FAILED"
    ws["H7"] = "BLOCKED"
    ws["I7"] = "Pass Rate"
    ws["J7"] = "Executable Pass Rate"

    style_header_row(
        ws,
        6,
        {
            2: yellow,
            3: grey,
            4: grey,
            5: grey,
            6: grey,
            7: grey,
            8: grey,
            9: grey,
            10: grey,
        },
        header_font,
        thin,
    )
    style_header_row(
        ws,
        7,
        {
            2: yellow,
            3: grey,
            4: grey,
            5: blue,
            6: grey,
            7: grey,
            8: grey,
            9: blue,
            10: blue,
        },
        header_font,
        thin,
    )
    for col in range(2, 11):
        ws.cell(6, col).border = thin
        ws.cell(7, col).border = thin
        ws.cell(6, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(7, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start = 8
    for idx, p in enumerate(plans):
        r = start + idx
        design = int(p.get("caseTotal") or 0)
        # 模板中 Review 通常等于 Design；MS 执行统计不含评审数，先按已设计=已评审
        review = design
        passed = int(p.get("successCount") or 0)
        failed = int(p.get("errorCount") or 0) + int(p.get("fakeErrorCount") or 0)
        blocked = int(p.get("blockCount") or 0) + int(p.get("pendingCount") or 0)

        ws.cell(r, 2, p.get("name") or "")
        ws.cell(r, 3, design)
        ws.cell(r, 4, review)
        ws.cell(r, 5, f"=IF(C{r}=0,0,D{r}/C{r})")
        ws.cell(r, 6, passed)
        ws.cell(r, 7, failed)
        ws.cell(r, 8, blocked)
        ws.cell(r, 9, f"=IF(C{r}=0,0,F{r}/C{r})")
        ws.cell(r, 10, f"=IF((F{r}+G{r})=0,0,F{r}/(F{r}+G{r}))")

        for col in range(2, 11):
            cell = ws.cell(r, col)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
        ws.cell(r, 5).number_format = "0%"
        ws.cell(r, 9).number_format = "0%"
        ws.cell(r, 10).number_format = "0%"

        # 100% 标绿（按数值预计算，公式打开后也会显示）
        for col in (5, 9, 10):
            # 预填显示色：若计算结果为 100%
            if col == 5 and design and review == design:
                ws.cell(r, col).font = green_font
            if col == 9 and design and passed == design:
                ws.cell(r, col).font = green_font
            if col == 10 and (passed + failed) > 0 and failed == 0 and passed > 0:
                ws.cell(r, col).font = green_font

    # 合计行
    end = start + len(plans) - 1 if plans else start
    total_row = end + 1
    ws.cell(total_row, 2, "Total")
    ws.cell(total_row, 2).font = Font(bold=True)
    if plans:
        for col, letter in [(3, "C"), (4, "D"), (6, "F"), (7, "G"), (8, "H")]:
            ws.cell(total_row, col, f"=SUM({letter}{start}:{letter}{end})")
        ws.cell(total_row, 5, f"=IF(C{total_row}=0,0,D{total_row}/C{total_row})")
        ws.cell(total_row, 9, f"=IF(C{total_row}=0,0,F{total_row}/C{total_row})")
        ws.cell(total_row, 10, f"=IF((F{total_row}+G{total_row})=0,0,F{total_row}/(F{total_row}+G{total_row}))")
        ws.cell(total_row, 5).number_format = "0%"
        ws.cell(total_row, 9).number_format = "0%"
        ws.cell(total_row, 10).number_format = "0%"
    for col in range(2, 11):
        cell = ws.cell(total_row, col)
        cell.border = thin
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")

    ws.column_dimensions["B"].width = 55
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 32
    ws.freeze_panes = "C8"
    return wb


def main() -> int:
    json_path = run_fetch()
    data, plans = load_plans(json_path)
    module_name = (data.get("module") or {}).get("name") or "OBIS-20260622-20260703"
    summary = data.get("summary") or {}

    wb = build_workbook(module_name, summary, plans)
    out_dir = ROOT / "exports" / "metersphere"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"{module_name}_child_plans_{stamp}.xlsx"
    wb.save(out_path)
    print(f"[PASS] Excel generated: {out_path}")
    print(f"       plans={len(plans)} source={json_path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
